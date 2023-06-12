from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import time

arquivo = load_workbook('List.xlsx')
aba = arquivo.active

options = ChromeOptions()
# options.add_argument("--headless")
options.add_argument("--log-level=3")

driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
driver.maximize_window()

lista = []

linha = 2

while aba.cell(linha, 2).value != None:
    e = aba.cell(linha, 1).value
    c = aba.cell(linha, 2).value
    cod = aba.cell(linha, 3).value
    lt = []
    lt.append(e)
    lt.append(c)
    lt.append(cod)
    lista.append(lt)
    linha += 1

wb = Workbook()
for i in lista:
    wb.create_sheet(i[1])

wb.create_sheet('UNIFICADA')

ab = wb['UNIFICADA']

ab.cell(1, 1).value = 'First Name'
ab.cell(1, 2).value = 'Last Name'
ab.cell(1, 3).value = 'E-mail'
ab.cell(1, 4).value = 'Phone'
ab.cell(1, 5).value = 'Cell'
ab.cell(1, 6).value = 'City'
ab.cell(1, 7).value = 'State'
ab.cell(1, 8).value = 'Empresa'
ab.cell(1, 9).value = 'Site'
ab.cell(1, 1).font = Font(bold=True)
ab.cell(1, 2).font = Font(bold=True)
ab.cell(1, 3).font = Font(bold=True)
ab.cell(1, 4).font = Font(bold=True)
ab.cell(1, 5).font = Font(bold=True)
ab.cell(1, 6).font = Font(bold=True)
ab.cell(1, 7).font = Font(bold=True)
ab.cell(1, 8).font = Font(bold=True)
ab.cell(1, 9).font = Font(bold=True)


for i in wb:
    if i.title == "Sheet":
        wb.remove(i)
    else:
        i.cell(1, 1).value = 'First Name'
        i.cell(1, 2).value = 'Last Name'
        i.cell(1, 3).value = 'E-mail'
        i.cell(1, 4).value = 'Phone'
        i.cell(1, 5).value = 'Cell'
        i.cell(1, 6).value = 'City'
        i.cell(1, 7).value = 'State'
        i.cell(1, 8).value = 'Empresa'
        i.cell(1, 9).value = 'Site'
        i.cell(1, 1).font = Font(bold=True)
        i.cell(1, 2).font = Font(bold=True)
        i.cell(1, 3).font = Font(bold=True)
        i.cell(1, 4).font = Font(bold=True)
        i.cell(1, 5).font = Font(bold=True)
        i.cell(1, 6).font = Font(bold=True)
        i.cell(1, 7).font = Font(bold=True)
        i.cell(1, 8).font = Font(bold=True)
        i.cell(1, 9).font = Font(bold=True)

wb.save('Result.xlsx')


def kw():
    empresa = 'Keller Williams'
    site = 'https://www.kw.com/'
    for i in lista:
        try:
            estado = i[0]
            cidade = i[1]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(f'https://www.kw.com/agent/search/{estado}/{cidade}/')
            time.sleep(5)
            total = driver.find_element(
                By.XPATH, '//*[@id="__next"]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/div').text
            lt_total = total.split(' ')
            max = int(lt_total[1].replace('.', ''))
            cont = 1150
            while cont <= max-1:
                print(cont)
                try:
                    while True:
                        try:
                            driver.find_elements(By.CLASS_NAME, 'FindAgentRoute__row')[
                                cont].click()
                            break
                        except:
                            driver.execute_script(
                                "window.scroll(0,100000000000);")
                            time.sleep(2)
                    time.sleep(8)
                    nome = driver.find_element(
                        By.CLASS_NAME, 'AgentContent__name').text
                    n = nome.split(' ')
                    try:
                        email = driver.find_element(
                            By.CLASS_NAME, 'AgentInformation__factBody').text
                    except:
                        email = ''
                    try:
                        cel = driver.find_element(
                            By.CLASS_NAME, 'AgentInformation__phoneMobileNumber').text
                    except:
                        cel = ''
                    try:
                        fixo = driver.find_element(
                            By.CLASS_NAME, 'AgentInformation__phoneOfficeNumber').text
                    except:
                        fixo = ''
                    aba.cell(max_linha, 1).value = n[0]
                    aba.cell(max_linha, 2).value = n[-1]
                    aba.cell(max_linha, 3).value = email
                    aba.cell(max_linha, 4).value = fixo
                    aba.cell(max_linha, 5).value = cel
                    aba.cell(max_linha, 6).value = cidade
                    aba.cell(max_linha, 7).value = estado
                    aba.cell(max_linha, 8).value = empresa
                    aba.cell(max_linha, 9).value = site
                    max_linha += 1
                    cont += 1
                    wb.save('Result.xlsx')
                    driver.get(
                        f'https://www.kw.com/agent/search/{estado}/{cidade}/')
                    time.sleep(5)
                except:
                    driver.get(
                        f'https://www.kw.com/agent/search/{estado}/{cidade}/')
                    time.sleep(5)
                    cont += 1
        except:
            continue


def remax():
    empresa = 'Remax'
    site = 'https://www.remax.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            est = i[0]
            cid = i[1]
            aba = wb[cidade]
            cidade = cidade.lower()
            estado = estado.lower()
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(
                f'https://www.remax.com/real-estate-agents/{cidade}-{estado}')
            time.sleep(30)
            lista_links = []
            cont = 2
            while True:
                agents = driver.find_elements(By.CLASS_NAME, 'd-agent-card')
                time.sleep(2)
                if len(agents) == 0:
                    break
                if agents[0].find_element(By.TAG_NAME, 'a').get_attribute('href') in lista_links:
                    break
                for i in agents:
                    link = i.find_element(By.TAG_NAME, 'a')
                    lk = link.get_attribute('href')
                    lista_links.append(lk)
                time.sleep(1)
                driver.execute_script("window.scroll(0,1000000);")
                time.sleep(2)
                try:
                    source = driver.find_element(
                        By.XPATH, '//*[@id="__layout"]/div/main/div/form/div/div[2]/div[3]/div/div[3]/input')
                    action = ActionChains(driver)
                    action.double_click(source)
                    action.perform()
                    action.send_keys(cont)
                    action.perform()
                    time.sleep(1)
                    cont += 1
                    time.sleep(10)
                except:
                    pass
            for i in lista_links:
                try:
                    time.sleep(2)
                    driver.get(i)
                    time.sleep(2)
                    nome = driver.find_element(By.CLASS_NAME, 'mt-6').text
                    n = nome.split(' ')
                    try:
                        elemento = driver.find_element(
                            By.CLASS_NAME, 'bio-phone')
                        try:
                            fixo = elemento.find_elements(
                                By.CLASS_NAME, 'phone-link')[0].text
                            cel = elemento.find_elements(
                                By.CLASS_NAME, 'phone-link')[1].text
                        except:
                            fixo = ""
                            cel = elemento.find_element(
                                By.CLASS_NAME, 'phone-link').text
                    except:
                        fixo = ''
                        cel = ''
                    aba.cell(max_linha, 1).value = n[0]
                    aba.cell(max_linha, 2).value = n[-1]
                    aba.cell(max_linha, 4).value = fixo
                    aba.cell(max_linha, 5).value = cel
                    aba.cell(max_linha, 6).value = cid
                    aba.cell(max_linha, 7).value = est
                    aba.cell(max_linha, 8).value = empresa
                    aba.cell(max_linha, 9).value = site
                    ab.cell(m, 1).value = n[0]
                    ab.cell(m, 2).value = n[-1]
                    ab.cell(m, 4).value = fixo
                    ab.cell(m, 5).value = cel
                    ab.cell(m, 6).value = cid
                    ab.cell(m, 7).value = est
                    ab.cell(m, 8).value = empresa
                    ab.cell(m, 9).value = site
                    max_linha += 1
                    m += 1
                    wb.save('Result.xlsx')
                except:
                    continue
        except:
            continue


def coldwellbanker():
    empresa = 'Coldwell Banker'
    site = 'https://www.coldwellbanker.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(
                f'https://www.coldwellbanker.com/real-estate-agents?companyId=reset')
            time.sleep(3)
            driver.find_element(By.ID, 'agentSearch').send_keys(
                cidade, ', ', estado, Keys.DOWN, Keys.ENTER)
            time.sleep(500)
            agents = driver.find_elements(By.CLASS_NAME, 'results-row')
            for i in agents:
                try:
                    nome = i.find_element(
                        By.CLASS_NAME, 'heading-std-mobile').text
                    try:
                        cel = i.find_element(
                            By.CLASS_NAME, 'agent-contact-phone').text
                    except:
                        cel = ''
                    n = nome.split(' ')
                    aba.cell(max_linha, 1).value = n[0]
                    aba.cell(max_linha, 2).value = n[-1]
                    aba.cell(max_linha, 5).value = cel
                    aba.cell(max_linha, 6).value = cidade
                    aba.cell(max_linha, 7).value = estado
                    aba.cell(max_linha, 8).value = empresa
                    aba.cell(max_linha, 9).value = site
                    ab.cell(m, 1).value = n[0]
                    ab.cell(m, 2).value = n[-1]
                    ab.cell(m, 5).value = cel
                    ab.cell(m, 6).value = cidade
                    ab.cell(m, 7).value = estado
                    ab.cell(m, 8).value = empresa
                    ab.cell(m, 9).value = site
                    max_linha += 1
                    m += 1
                    wb.save('Result.xlsx')
                except:
                    continue
        except:
            continue


def century21():
    empresa = 'Century 21'
    site = 'https://www.century21.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            c1 = cidade.replace(' ', '-')
            c2 = cidade.replace(' ', '')
            driver.get(
                f'https://www.century21.com/real-estate-agents/{c1.lower()}-{estado.lower()}/LC{estado.upper()}{c2.upper()}/')
            time.sleep(3)
            cont = 12
            while True:
                total = driver.find_element(
                    By.XPATH, '//*[@id="content"]/div[3]/div/div[1]').text
                t = total.split(' ')
                if int(t[1].replace(',', '')) > int(t[-1].replace(',', '')):
                    break
                agents = driver.find_elements(By.CLASS_NAME, 'agent-card')
                for i in agents:
                    try:
                        nome = i.find_element(By.CLASS_NAME, 'card-title').text
                        n = nome.split(' ')
                        try:
                            cel = i.find_element(
                                By.CLASS_NAME, 'card__inner-link').text
                        except:
                            cel = ''
                        aba.cell(max_linha, 1).value = n[0]
                        aba.cell(max_linha, 2).value = n[-1]
                        aba.cell(max_linha, 5).value = cel
                        aba.cell(max_linha, 6).value = cidade
                        aba.cell(max_linha, 7).value = estado
                        aba.cell(max_linha, 8).value = empresa
                        aba.cell(max_linha, 9).value = site
                        ab.cell(m, 1).value = n[0]
                        ab.cell(m, 2).value = n[-1]
                        ab.cell(m, 5).value = cel
                        ab.cell(m, 6).value = cidade
                        ab.cell(m, 7).value = estado
                        ab.cell(m, 8).value = empresa
                        ab.cell(m, 9).value = site
                        m += 1
                        max_linha += 1
                        wb.save('Result.xlsx')
                    except:
                        continue
                time.sleep(2)
                driver.get(
                    f'https://www.century21.com/real-estate-agents/{c1.lower()}-{estado.lower()}/LC{estado.upper()}{c2.upper()}/?s={str(cont)}')
                cont += 12
                time.sleep(3)
        except:
            continue


def bhhs():
    empresa = 'Berkshire Hathaway'
    site = 'https://www.bhhs.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(
                f'https://www.bhhs.com/agent-search-results?city={cidade}%2C%2B{estado}%2C%2BUnited%2BStates')
            time.sleep(10)
            try:
                driver.find_element(
                    By.ID, 'adroll_reject').click()
                time.sleep(3)
            except:
                pass
            driver.find_element(
                By.XPATH, '//*[@id="ae-skip-to-content"]/div/div/div/section/section[2]/div/div/div[2]/div/div[2]/section/div/div').click()
            time.sleep(2)
            driver.find_element(
                By.XPATH, '//*[@id="ae-skip-to-content"]/div/div/div/section/section[2]/div/div/div[2]/div/div[2]/section/div/ul/li[5]').click()
            time.sleep(10)
            lista_links = []
            lt_check = []
            while True:
                agents = driver.find_elements(
                    By.CLASS_NAME, 'cmp-agent-results-list-view')
                verifica = agents[0].find_element(
                    By.CLASS_NAME, 'associate__name').text
                if verifica in lt_check:
                    break
                else:
                    lt_check.append(verifica)
                for i in agents:
                    link = i.find_element(
                        By.CLASS_NAME, 'btn-secondary').get_attribute('href')
                    lista_links.append(link)
                driver.find_element(
                    By.CLASS_NAME, 'cmp-search-results-pagination__arrow--next').click()
                time.sleep(20)
            for i in lista_links:
                try:
                    time.sleep(1)
                    driver.get(i)
                    time.sleep(5)
                    try:
                        driver.find_element(
                            By.ID, 'adroll_reject').click()
                        time.sleep(3)
                    except:
                        pass
                    nome = driver.find_element(
                        By.CLASS_NAME, 'homepage_link').text
                    n = nome.split(' ')
                    try:
                        email = driver.find_element(
                            By.CLASS_NAME, 'cmp-agent-details__mail').text
                    except:
                        email = ''
                    try:
                        cel = driver.find_element(
                            By.CLASS_NAME, 'cmp-agent-details__phone-number').text
                    except:
                        cel = ''
                    aba.cell(max_linha, 1).value = n[0]
                    aba.cell(max_linha, 2).value = n[-1]
                    aba.cell(max_linha, 3).value = email
                    aba.cell(max_linha, 5).value = cel
                    aba.cell(max_linha, 6).value = cidade
                    aba.cell(max_linha, 7).value = estado
                    aba.cell(max_linha, 8).value = empresa
                    aba.cell(max_linha, 9).value = site
                    ab.cell(m, 1).value = n[0]
                    ab.cell(m, 2).value = n[-1]
                    ab.cell(m, 3).value = email
                    ab.cell(m, 5).value = cel
                    ab.cell(m, 6).value = cidade
                    ab.cell(m, 7).value = estado
                    ab.cell(m, 8).value = empresa
                    ab.cell(m, 9).value = site
                    m += 1
                    max_linha += 1
                    wb.save('Result.xlsx')
                except:
                    continue
        except:
            continue


def compass():
    empresa = 'Compass'
    site = 'https://www.compass.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            codigo = i[2]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(
                f'https://www.compass.com/agents/locations/{cidade.lower()}-{estado.lower()}/{codigo}/')
            time.sleep(3)
            lt_verifica = []
            while True:
                agents = driver.find_elements(By.CLASS_NAME, 'agentCard')
                verifica = agents[0].find_element(
                    By.CLASS_NAME, 'agentCard-name').text
                if verifica in lt_verifica:
                    break
                else:
                    lt_verifica.append(verifica)
                for i in agents:
                    try:
                        nome = i.find_element(
                            By.CLASS_NAME, 'agentCard-name').text
                        n = nome.split(' ')
                        try:
                            email = i.find_element(
                                By.CLASS_NAME, 'agentCard-email').text
                        except:
                            email = ''
                        try:
                            cel = i.find_element(
                                By.CLASS_NAME, 'agentCard-phone').text
                        except:
                            cel = ''
                        aba.cell(max_linha, 1).value = n[0]
                        aba.cell(max_linha, 2).value = n[-1]
                        aba.cell(max_linha, 3).value = email
                        aba.cell(max_linha, 5).value = cel.replace('M: ', '')
                        aba.cell(max_linha, 6).value = cidade
                        aba.cell(max_linha, 7).value = estado
                        aba.cell(max_linha, 8).value = empresa
                        aba.cell(max_linha, 9).value = site
                        ab.cell(m, 1).value = n[0]
                        ab.cell(m, 2).value = n[-1]
                        ab.cell(m, 3).value = email
                        ab.cell(m, 5).value = cel.replace('M: ', '')
                        ab.cell(m, 6).value = cidade
                        ab.cell(m, 7).value = estado
                        ab.cell(m, 8).value = empresa
                        ab.cell(m, 9).value = site
                        max_linha += 1
                        m += 1
                        wb.save('Result.xlsx')
                    except:
                        continue
                driver.find_element(
                    By.CLASS_NAME, 'cx-react-pagination-next').click()
                time.sleep(5)
        except:
            continue


def exprealty():
    empresa = 'Exprealty'
    site = 'https://exprealty.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            codigo = i[2]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(
                f'https://exprealty.com/agents/#/?city={cidade}%2C+{estado}&country=US')
            time.sleep(3)
            lt_verifica = []
            while True:
                agents = driver.find_elements(By.CLASS_NAME, 'Agents_grid__3ia8J')
                for i in agents:
                    try:
                        nome = i.find_element(
                            By.CLASS_NAME, 'PersonCard_name__3WLGO').text
                        n = nome.split(' ')
                        try:
                            email = i.find_element(
                                By.CLASS_NAME, 'PersonCard_point__j-l3P').text
                        except:
                            email = ''
                        try:
                            cel = i.find_element(
                                By.CLASS_NAME, 'PersonCard_point__j-l3P').text
                        except:
                            cel = ''
                        aba.cell(max_linha, 1).value = n[0]
                        aba.cell(max_linha, 2).value = n[-1]
                        aba.cell(max_linha, 3).value = email
                        aba.cell(max_linha, 5).value = cel.replace('M: ', '')
                        aba.cell(max_linha, 6).value = cidade
                        aba.cell(max_linha, 7).value = estado
                        aba.cell(max_linha, 8).value = empresa
                        aba.cell(max_linha, 9).value = site
                        ab.cell(m, 1).value = n[0]
                        ab.cell(m, 2).value = n[-1]
                        ab.cell(m, 3).value = email
                        ab.cell(m, 5).value = cel.replace('M: ', '')
                        ab.cell(m, 6).value = cidade
                        ab.cell(m, 7).value = estado
                        ab.cell(m, 8).value = empresa
                        ab.cell(m, 9).value = site
                        max_linha += 1
                        m += 1
                        wb.save('Result.xlsx')
                    except:
                        continue
                driver.find_element(
                    By.XPATH, '//*[@id="sidx-agent-directory"]/div[2]/div[3]/button[4]').click()
                time.sleep(5)
        except:
            continue

def exitrealty():
    empresa = 'Exitrealty'
    site = 'https://exitrealty.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            codigo = i[2]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(
                f'https://exitrealty.com/agents/area-{cidade}__{estado}__United_States')
            time.sleep(3)
            while True:
                agents = driver.find_elements(By.CLASS_NAME, 'styles__RenderAgentsResult-sc-fr1lq0-24 ekORmZ')
                for i in agents:
                    try:
                        nome = i.find_element(
                            By.CLASS_NAME, 'styles__AgentName-sc-fr1lq0-13 bShrxP').text
                        n = nome.split(' ')
                        try:
                            email = i.find_element(
                                By.ID, 'agent-search-email-0').text
                        except:
                            email = ''
                        try:
                            cel = i.find_element(
                                By.CLASS_NAME, 'styles__ContactValue-sc-fr1lq0-19 cVPHAv').text
                        except:
                            cel = ''
                        aba.cell(max_linha, 1).value = n[0]
                        aba.cell(max_linha, 2).value = n[-1]
                        aba.cell(max_linha, 3).value = email
                        aba.cell(max_linha, 5).value = cel
                        aba.cell(max_linha, 6).value = cidade
                        aba.cell(max_linha, 7).value = estado
                        aba.cell(max_linha, 8).value = empresa
                        aba.cell(max_linha, 9).value = site
                        ab.cell(m, 1).value = n[0]
                        ab.cell(m, 2).value = n[-1]
                        ab.cell(m, 3).value = email
                        ab.cell(m, 5).value = cel
                        ab.cell(m, 6).value = cidade
                        ab.cell(m, 7).value = estado
                        ab.cell(m, 8).value = empresa
                        ab.cell(m, 9).value = site
                        max_linha += 1
                        m += 1
                        wb.save('Result.xlsx')
                    except:
                        continue
                driver.find_element(
                    By.XPATH, '//*[@id="outer-container"]/div[6]/div[3]/div/button[4]').click()
                time.sleep(5)
        except:
            continue

def sothersby():
    empresa = 'Sothersby'
    site = 'https://www.sothebysrealty.com/'
    for i in lista:
        try:
            ab = wb['UNIFICADA']
            for x in range(1, 200000):
                if ab.cell(row=x, column=1).value == None:
                    m = x
                    break
            estado = i[0]
            cidade = i[1]
            aba = wb[cidade]
            for x in range(1, 200000):
                if aba.cell(row=x, column=2).value == None:
                    max_linha = x
                    break
            driver.get(f'https://www.sothebysrealty.com/eng/associates/washington-dc-usa')
            n = driver.find_element(By.XPATH, '//*[contains(text(), "results") and @class="h6"]').text.split(' ')[0]
            num = int(n)
            time.sleep(3)
            num_page = 1
            pages = []
            while True:
                agents = driver.find_elements(By.CLASS_NAME, 'Entities-card__cta')
                #print(agents)
                for i in agents:
                        if i.is_displayed() == False:
                            continue
                        else:
                            pages.append(i.get_attribute('href'))
                print(len(pages))
                if len(pages) < num:
                    num_page += 1
                    driver.get(f'https://www.sothebysrealty.com/eng/associates/washington-dc-usa/{num_page}-pg')
                    time.sleep(5)
                else:
                    break
            for p in pages:
                    driver.get(p)
                    time.sleep(3)
                    try:
                            nome = driver.find_element(By.CLASS_NAME, 'u-color-dark-blue').text
                            print(nome)
                            n = nome.split(' ')
                            try:
                                email = driver.find_element(By.CLASS_NAME, 'GetInTouch__agent-emailvCard').text
                                print(email)
                            except:
                                email = ''
                            try:
                                cel = driver.find_element(By.CLASS_NAME, 'EntityPhones__wrapper').text
                                print(cel)
                            except:
                                cel = ''
                            aba.cell(max_linha, 1).value = n[0]
                            aba.cell(max_linha, 2).value = n[-1]
                            aba.cell(max_linha, 3).value = email
                            aba.cell(max_linha, 5).value = cel
                            aba.cell(max_linha, 6).value = cidade
                            aba.cell(max_linha, 7).value = estado
                            aba.cell(max_linha, 8).value = empresa
                            aba.cell(max_linha, 9).value = site
                            ab.cell(m, 1).value = n[0]
                            ab.cell(m, 2).value = n[-1]
                            ab.cell(m, 3).value = email
                            ab.cell(m, 5).value = cel
                            ab.cell(m, 6).value = cidade
                            ab.cell(m, 7).value = estado
                            ab.cell(m, 8).value = empresa
                            ab.cell(m, 9).value = site
                            max_linha += 1
                            m += 1
                            wb.save('Result.xlsx')
                    except:
                        continue
        except:
            continue




print('''

Selecione o site para busca:

1 Keller Williams
2 ReMax
3 Coldwell Banker
4 Century 21
5 Berkshire Hathaway
6 Compass
7 exprealty
8 exitrealty
9 sothersby
10 todos
    
''')

while True:
    op = input('Selecione sua opção e aperte ENTER: ')
    if op == str(1):
        kw()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(2):
        remax()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(3):
        coldwellbanker()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(4):
        century21()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(5):
        bhhs()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(6):
        compass()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(7):
        exprealty()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(8):
        exitrealty()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(9):
        sothersby()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    elif op == str(10):
        remax()
        coldwellbanker()
        century21()
        bhhs()
        compass()
        exprealty()
        exitrealty()
        sothersby()
        driver.quit()
        input('\nBusca Finalizada!')
        break
    else:
        print('Opção inválida, selecione novamente')
